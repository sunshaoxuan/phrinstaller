import openpyxl
import json

def extract_specific_sheets(path, sheets):
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for name in sheets:
        if name in wb.sheetnames:
            sheet = wb[name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                if any(row):
                    rows.append([str(cell) if cell is not None else "" for cell in row])
            result[name] = rows
    return result

if __name__ == "__main__":
    path = "c:\\workspace\\PHRInstaller\\docs\\環境構築手順書_ONPREMISE.xlsx"
    target_sheets = ['庶務事務服务手順', '４．OHR服务をインストール手順', '５ 業務検証', 'F１．MINIOを配置する']
    # Wait, the names might be slightly different in the terminal output.
    # Re-reading sheet names correctly.
    wb = openpyxl.load_workbook(path, read_only=True)
    print(f"Sheets available: {wb.sheetnames}")
    
    data = extract_specific_sheets(path, wb.sheetnames)
    for sname, sdata in data.items():
        print(f"--- {sname} ---")
        for row in sdata[:50]: # First 50 rows per sheet should be enough for steps
            print("\t".join(row))
