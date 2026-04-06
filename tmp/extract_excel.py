import openpyxl
import json
import sys

def extract_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            if any(row): # Skip empty rows
                rows.append([str(cell) if cell is not None else "" for cell in row])
        result[sheet_name] = rows
    return result

if __name__ == "__main__":
    try:
        excel_path = "c:\\workspace\\PHRInstaller\\docs\\環境構築手順書_ONPREMISE.xlsx"
        data = extract_excel(excel_path)
        print(json.dumps(data, ensure_ascii=False, indent=2))
    except Exception as e:
        print(f"Error: {e}")
