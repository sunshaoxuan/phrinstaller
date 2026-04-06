import openpyxl
import os

path = "c:\\workspace\\PHRInstaller\\docs\\環境構築手順書_ONPREMISE.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
os.makedirs("c:\\workspace\\PHRInstaller\\tmp\\excel_extracts", exist_ok=True)

for name in wb.sheetnames:
    sheet = wb[name]
    output_pf = os.path.join("c:\\workspace\\PHRInstaller\\tmp\\excel_extracts", f"{name}.txt")
    with open(output_pf, "w", encoding="utf-8") as f:
        for row in sheet.iter_rows(values_only=True):
            if any(row):
                f.write("\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n")
    print(f"Saved {name} to {output_pf}")
